#!/usr/bin/env python3
"""Genera catalogo-metricas-haulmer.xlsx con diseño desde el HTML del playbook."""
import json
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HTML = Path(__file__).resolve().parent.parent / "index.html"
OUT = Path(__file__).resolve().parent.parent / "catalogo-metricas-haulmer.xlsx"
EXTRACT = Path(__file__).resolve().parent / "extract_catalog_rows.mjs"

TYPE_STYLE = {
    "ns": ("EEEDFE", "26215C"),
    "clave": ("E3F2FD", "0D47A1"),
    "driver": ("FFF8E7", "412402"),
    "salud": ("E0F7FA", "006064"),
    "operativa": ("F1EFE8", "5F5E5A"),
}

HEADERS = [
    "Unidad",
    "Nodo",
    "Tipo",
    "Métrica",
    "Proceso de maduración",
    "Descripción",
    "Qué preguntas responde",
    "Área dueña",
    "Fuente / tabla",
    "Fórmula",
]

COL_WIDTHS = [12, 14, 18, 34, 30, 42, 34, 12, 28, 36]


def thin_border():
    s = Side(style="thin", color="E8E6DC")
    return Border(left=s, right=s, top=s, bottom=s)


def load_rows():
    proc = subprocess.run(
        ["node", str(EXTRACT), str(HTML.resolve())],
        capture_output=True,
        text=True,
        check=True,
    )
    if proc.stderr:
        print(proc.stderr, file=__import__("sys").stderr)
    return json.loads(proc.stdout)


def build_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo métricas"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "Catálogo de métricas — Haulmer"
    t.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1B1B18")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:J2")
    s = ws["A2"]
    s.value = f"Generado {datetime.now().strftime('%d/%m/%Y %H:%M')} · {len(rows)} métricas · CS mostrado como CX"
    s.font = Font(name="Calibri", size=10, italic=True, color="5F5E5A")
    s.fill = PatternFill("solid", fgColor="FAFAF8")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 3
    header_fill = PatternFill("solid", fgColor="26215C")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[header_row].height = 28

    data_start = header_row + 1
    zebra_a = PatternFill("solid", fgColor="FFFFFF")
    zebra_b = PatternFill("solid", fgColor="FAFAF8")
    base_font = Font(name="Calibri", size=10, color="1B1B18")
    empty_font = Font(name="Calibri", size=10, italic=True, color="A09E97")

    for i, row in enumerate(rows):
        r = data_start + i
        values = [
            row["unit"],
            row["runLabel"],
            row["typeLabel"],
            row["name"],
            row.get("proceso") or "",
            row["descripcion"] or "Por definir",
            row["preguntas"] or "Por definir",
            row["area"] or "Por definir",
            row["fuente"] or "Por definir con el equipo",
            row["formula"] or "Por definir",
        ]
        row_fill = zebra_b if i % 2 else zebra_a
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 3:
                bg, fg = TYPE_STYLE.get(row["type"], ("F1EFE8", "5F5E5A"))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif col == 4:
                cell.font = Font(name="Calibri", size=10, bold=True, color="1B1B18")
                cell.fill = row_fill
            elif col == 5:
                cell.fill = row_fill
                if val:
                    cell.font = base_font
                else:
                    cell.value = ""
            else:
                cell.fill = row_fill
                if isinstance(val, str) and val.startswith("Por definir"):
                    cell.font = empty_font
                else:
                    cell.font = base_font
        ws.row_dimensions[r].height = 42

    ws.freeze_panes = ws.cell(row=data_start, column=1)
    ws.auto_filter.ref = f"A{header_row}:J{data_start + len(rows) - 1}"

    for idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(OUT)
    print(f"OK {len(rows)} filas -> {OUT}")


if __name__ == "__main__":
    build_xlsx(load_rows())
